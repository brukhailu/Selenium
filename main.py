from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from flask import Flask, request, render_template, redirect, url_for
from openpyxl import load_workbook
import os

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}
def login_user(username, password):
    print("Logged In")

def register_agent(agent_data):
    # Initialize the WebDriver service
    service = Service(executable_path="chromedriver.exe")
    driver = webdriver.Chrome(service=service)

    try:
        # Open the registration URL
        driver.get("http://localhost:5002/")
        login_user()
        print("URL opened successfully")

        # Fill in the registration form
        driver.find_element(By.ID, 'name').send_keys(agent_data[0])
        time.sleep(3)
        driver.find_element(By.ID, 'email').send_keys(agent_data[1])
        time.sleep(3)
        driver.find_element(By.ID, 'phone').send_keys(agent_data[2])
        time.sleep(3)
        driver.find_element(By.ID, 'address').send_keys(agent_data[3])
        time.sleep(3)

        print("Form filled successfully")

        # Submit the form
        driver.find_element(By.ID, 'submit').click()
        print("Form submitted successfully")

        # Wait for the form submission to process
        time.sleep(15)

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Close the WebDriver
        driver.quit()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


@app.route('/')
def upload_form():
    return '''
    <!doctype html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Upload Excel File</title>
    </head>
    <body>
        <h1>Upload Excel File</h1>
        <form method="post" action="/upload" enctype="multipart/form-data">
            <input type="file" name="file">
            <input type="submit" value="Upload">
        </form>
    </body>
    </html>
    '''


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return 'No file part'
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'
    if file and allowed_file(file.filename):
        filename = file.filename
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        if filename.rsplit('.', 1)[1].lower() == 'xlsx':
            # Process xlsx file
            workbook = load_workbook(file_path)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
        elif filename.rsplit('.', 1)[1].lower() == 'xls':
            # Process xls file
            import xlrd
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            data = []
            for row in range(sheet.nrows):
                data.append(sheet.row_values(row))
        else:
            return 'Invalid file type'

        for agent_data in data:
            if len(agent_data) >= 4:
                print(f"Registering agent: {agent_data}")
                register_agent(agent_data)

        return f'File uploaded and agents registered: {data}'
    return 'Invalid file type'


if __name__ == '__main__':
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    app.run(host='localhost', port=5000, debug=True)
